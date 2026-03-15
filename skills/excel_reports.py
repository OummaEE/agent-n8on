"""
Skill: excel_reports
Description: Create advanced Excel reports with charts, formatting, and multiple sheets.
Requires: openpyxl (already installed via install.bat)
Author: Jane's Agent Builder
"""

SKILL_NAME = "excel_reports"
SKILL_VERSION = "1.0"
SKILL_DESCRIPTION = "Create Excel reports with charts, formatting, multiple sheets, and data analysis"
SKILL_TOOLS = {
    "excel_create_report": {
        "description": "Create a formatted Excel report from data",
        "args": {
            "path": "Path to save the Excel file",
            "title": "Report title",
            "headers": "List of column headers",
            "data": "List of rows (each row is a list of values)",
            "chart_type": "Optional: bar, line, pie (default: none)"
        },
        "example": '{"tool": "excel_create_report", "args": {"path": "C:/Users/Dator/Desktop/report.xlsx", "title": "Sales Report", "headers": ["Month", "Revenue", "Costs"], "data": [["Jan", 1000, 500], ["Feb", 1200, 600]], "chart_type": "bar"}}'
    },
    "excel_read": {
        "description": "Read an Excel file and return its contents as text",
        "args": {"path": "Path to the Excel file"},
        "example": '{"tool": "excel_read", "args": {"path": "C:/Users/Dator/Desktop/data.xlsx"}}'
    },
    "excel_analyze": {
        "description": "Analyze an Excel file — show stats, columns, row counts",
        "args": {"path": "Path to the Excel file"},
        "example": '{"tool": "excel_analyze", "args": {"path": "C:/Users/Dator/Desktop/data.xlsx"}}'
    }
}


def excel_create_report(path: str, title: str, headers: list, data: list,
                        chart_type: str = "") -> str:
    """Create a formatted Excel report"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31] if title else "Report"

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        title_font = Font(bold=True, size=16, color="2F5496")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center')

        # Date
        import datetime
        ws.cell(row=2, column=1, value=f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

        # Headers (row 4)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_len = len(str(headers[col - 1])) if col <= len(headers) else 10
            for row in range(5, len(data) + 5):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

        # Chart (if requested)
        if chart_type and len(data) > 0 and len(headers) > 1:
            try:
                from openpyxl.chart import BarChart, LineChart, PieChart, Reference

                chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
                ChartClass = chart_classes.get(chart_type.lower(), BarChart)
                chart = ChartClass()
                chart.title = title
                chart.style = 10

                # Data reference (columns 2+ are numeric data)
                data_ref = Reference(ws, min_col=2, max_col=len(headers),
                                     min_row=4, max_row=len(data) + 4)
                cats = Reference(ws, min_col=1, min_row=5, max_row=len(data) + 4)

                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)
                chart.width = 20
                chart.height = 12

                ws.add_chart(chart, f"A{len(data) + 7}")
            except Exception as e:
                pass  # Chart is optional

        # Summary row
        summary_row = len(data) + 5
        ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, len(headers) + 1):
            # Check if column is numeric
            if data and isinstance(data[0][col - 1] if col - 1 < len(data[0]) else None, (int, float)):
                total = sum(row[col - 1] for row in data if col - 1 < len(row) and isinstance(row[col - 1], (int, float)))
                cell = ws.cell(row=summary_row, column=col, value=total)
                cell.font = Font(bold=True)
                cell.border = border

        # Save
        os.makedirs(os.path.dirname(path) if os.path.dirname(path) else '.', exist_ok=True)
        wb.save(path)
        return f"Excel report created: {path}\nSheets: 1, Rows: {len(data)}, Columns: {len(headers)}"

    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return f"Error creating Excel report: {str(e)}"


def excel_read(path: str) -> str:
    """Read an Excel file"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        
        lines = [f"=== {os.path.basename(path)} ==="]
        lines.append(f"Sheets: {', '.join(wb.sheetnames)}\n")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"--- Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ---")
            
            # Read first 50 rows
            for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row or 1), values_only=True):
                values = [str(v) if v is not None else "" for v in row]
                lines.append(" | ".join(values))
            
            if (ws.max_row or 0) > 50:
                lines.append(f"... ({ws.max_row - 50} more rows)")
            lines.append("")
        
        return "\n".join(lines)
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return f"Error reading Excel: {str(e)}"


def excel_analyze(path: str) -> str:
    """Analyze an Excel file"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        
        lines = [f"=== Analysis: {os.path.basename(path)} ==="]
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"\n📊 Sheet: {sheet_name}")
            lines.append(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
            
            if ws.max_row and ws.max_row > 1:
                # Get headers
                headers = [str(cell.value) for cell in ws[1] if cell.value]
                lines.append(f"  Headers: {', '.join(headers)}")
                
                # Analyze each column
                for col_idx, header in enumerate(headers, 1):
                    values = []
                    for row in range(2, min(ws.max_row + 1, 1000)):
                        v = ws.cell(row=row, column=col_idx).value
                        if v is not None:
                            values.append(v)
                    
                    if values and all(isinstance(v, (int, float)) for v in values):
                        lines.append(f"  {header}: min={min(values)}, max={max(values)}, "
                                     f"avg={sum(values)/len(values):.2f}, count={len(values)}")
                    else:
                        unique = len(set(str(v) for v in values))
                        lines.append(f"  {header}: {len(values)} values, {unique} unique")
        
        return "\n".join(lines)
    except ImportError:
        return "Error: openpyxl not installed. Run: pip install openpyxl"
    except Exception as e:
        return f"Error analyzing Excel: {str(e)}"


import os

TOOLS = {
    "excel_create_report": lambda args: excel_create_report(
        args.get("path", ""), args.get("title", "Report"),
        args.get("headers", []), args.get("data", []),
        args.get("chart_type", "")
    ),
    "excel_read": lambda args: excel_read(args.get("path", "")),
    "excel_analyze": lambda args: excel_analyze(args.get("path", "")),
}
